import pandas as pd
import requests
import time
import random
from openpyxl import load_workbook
from datetime import datetime


#
#
#

company_id = "123456"  # id компании, можно посмотреть на странице компании (например https://www.ozon.ru/seller/ooo-mebelnaya-fabrika-volzhanka-1234/products/?miniapp=seller_1234 - id компании 1234)
date_to = "2025-01-11"  # будет отправлять запросы пока не встретит указанную дату (лучше указывать нужную дату +1)
cookie = "cookie"  # Актуальные куки
cases = 500000  # Количество вопросов, которое нужно получить. 1 кейс = 10 вопросов

#
#
#


# Функция для получения вопросов
def get_questions(value_cases: int, cookie: str):
    global start_time
    start_time = time.time()  # Начинаем таймер

    url = "https://seller.ozon.ru/api/v1/question-list"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/115.0",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "ru",
        "Content-Type": "application/json",
        "x-o3-app-name": "seller-ui",
        "x-o3-language": "ru",
        "x-o3-company-id": company_id,
        "x-o3-page-type": "questions",
        "X-KL-Ajax-Request": "Ajax_Request",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "Cookie": cookie,
    }

    last_id = None
    last_timestamp = None

    # можно начать с определенной даты, указав конкретные id и timestamp:
    # last_id = "01-c9-76-aa-bc"
    # last_timestamp = "1728695750220301"

    data = {
        "sc_company_id": company_id,
        "with_brands": False,
        "with_counters": False,
        "company_type": "seller",
        "filter": {"status": "ALL"},
        "pagination_last_id": last_id,
        "last_published_at": last_timestamp,
    }

    # Цикл для получения указанного количества кейсов
    for cases in range(value_cases):
        success = False
        attempts = 0
        max_attempts = 60  # Количество попыток при ошибке

        while not success and attempts < max_attempts:
            try:
                response = requests.post(url, headers=headers, json=data)
                response.raise_for_status()
                result = response.json()["result"]
                data_at = ""

                # Проверка на дату
                for item in result:
                    data_at = item["published_at"].split("T")[0]
                    if item["published_at"].split("T")[0] == date_to:
                        print("Дата найдена, завершение выполнения.")
                        return  # Выход, если нужная дата найдена
                print(data_at)

                last_id = str(response.json()["pagination_last_id"])
                last_timestamp = str(response.json()["last_published_at"])
                data["pagination_last_id"] = last_id
                data["last_published_at"] = last_timestamp

                print(last_id)
                print(last_timestamp)
                print(f"Запрос {cases} из {value_cases}")

                # Сохраняем данные в файл
                save_xlsx(result)
                success = True
            except requests.exceptions.RequestException as e:
                print(
                    f"Ошибка: {e}. Повторная попытка {attempts + 1} из {max_attempts}."
                )
                attempts += 1
                time.sleep(random.uniform(70, 91))  # Задержка перед повтором
        if not success:
            print("Не удалось получить данные после 60 попыток.")
            break

        time.sleep(random.uniform(5, 6))  # Задержка перед следующим запросом
    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    print("Отчет готов!")


# Функция для сохранения данных в Excel
def save_xlsx(data):
    # Преобразуем данные в DataFrame
    df = pd.DataFrame(data)

    # Преобразуем даты в нужный формат
    df["published_at"] = df["published_at"].apply(
        lambda x: x.split("T")[0] + " " + x.split("T")[1].split(".")[0]
    )

    # Преобразуем информацию о бренде и авторе
    df["brand_info"] = df["brand_info"].apply(lambda x: x["name"])
    df["author"] = df["author"].apply(lambda x: x["name"])

    # Извлекаем дополнительные поля
    def extract_fields(cell):
        return {
            "Название товара": cell.get("title", ""),
            "Ссылка на товар": cell.get("url", ""),
            "Артикул товара": cell.get("offer_id", ""),
        }

    extracted = df["product"].apply(extract_fields).apply(pd.Series)
    df = pd.concat([df, extracted], axis=1)

    # Удаляем ненужные колонки
    df.drop("shareLink", axis=1, inplace=True)
    df.drop("company_info", axis=1, inplace=True)
    df.drop("is_answerable", axis=1, inplace=True)
    df.drop("usefulness_count", axis=1, inplace=True)
    df.drop("product", axis=1, inplace=True)

    # Переименовываем столбцы
    df.rename(
        columns={
            "id": "id",
            "sku": "Артикул OZON",
            "text": "Текст вопроса",
            "published_at": "Дата и время публикации вопроса",
            "author": "Имя автора",
            "brand_info": "Торговая сеть",
            "answers_total_count": "Ответов",
        },
        inplace=True,
    )

    # Получаем текущую дату для имени файла
    current_date = datetime.now().strftime("%Y-%m-%d")

    file_name = f"ozon_questions_{current_date}.xlsx"

    try:
        # Попытка загрузить существующий файл
        book = load_workbook(file_name)
        with pd.ExcelWriter(
            file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:
            # Находим первую пустую строку для записи данных
            startrow = writer.sheets["Sheet1"].max_row
            # Записываем данные в файл
            df.to_excel(
                writer,
                index=False,
                header=False,
                startrow=startrow,
                sheet_name="Sheet1",
            )
    except FileNotFoundError:
        # Если файл не найден, создаем новый
        df.to_excel(file_name, index=False)
        print("------------------------------------------------")
        print("Файл успешно создан!")
    print("Файл успешно обновлен!")
    end_time = time.time()
    execution_time = end_time - start_time
    execution_time_formatted = time.strftime("%H:%M:%S", time.gmtime(execution_time))
    print(f"Время выполнения программы: {execution_time_formatted}")
    print("------------------------------------------------")


if __name__ == "__main__":
    get_questions(cases, cookie)
